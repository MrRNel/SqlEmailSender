import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime

def export_data_to_excel(df, excel_file_path):
    try:
        # Create a new Excel workbook and add a worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"

        # Get the current date and time
        current_date = datetime.now().strftime("%Y-%m-%d_%H:%M")

         # Check if the DataFrame is empty
        if df.empty:
            # Merge cells and insert the message for no refuels
            ws.merge_cells('A1:E5')  # Merge cells for the message
            merged_cell = ws.cell(row=1, column=1, value=f'There were no loads on {current_date}.')
            merged_cell.font = Font(bold=True, color="FF0000")  # Red text
            merged_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Center text both horizontally and vertically
        else:
            # Insert 6 empty rows and keep track of the row where the table will start
            empty_rows = 6
            table_start_row = 1 + empty_rows

            for _ in range(empty_rows):
                ws.append([])

            # Convert the Pandas DataFrame to a list of lists
            data = dataframe_to_rows(df, index=False, header=True)

            # Write the data to the worksheet starting from the row where the table starts
            for r_idx, row in enumerate(data, table_start_row):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            # Create an Excel table from the data
            table = Table(displayName="DataTable", ref=ws.dimensions)
            style = TableStyleInfo(
                name="TableStyleMedium9", showFirstColumn=False,
                showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            table.tableStyleInfo = style
            ws.add_table(table)

            # Auto-size columns to fit their content
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
             # Specify a custom width for the date column (adjust as needed)
            date_column_width = 12  # You can adjust this width to your preference
            ws.column_dimensions['B'].width = date_column_width

            # Merge cells for the heading (3 rows)
            ws.merge_cells(f'A1:G3')  # Merge cells for the heading

            # Set formatting for the heading cell
            heading_cell = ws.cell(row=1, column=1, value=f'Daily Load Report ({current_date})')
            heading_cell.font = Font(bold=True)
            heading_cell.alignment = Alignment(horizontal='center', vertical='center')  # Center text both horizontally and vertically
            heading_cell.fill = PatternFill(start_color='538DD5', end_color='538DD5', fill_type='solid')

        # Define the Excel file name with client name and date
        excel_file_name = f'Load_Report_{current_date}.xlsx'

        excel_file_path_string = os.path.join(excel_file_path, excel_file_name)

            # Save the Excel workbook to the specified path
        wb.save(excel_file_path_string)
            
        return excel_file_path_string  # Return the file path where the Excel file was saved

    except Exception as e:
        print(f"Error exporting data to Excel: {str(e)}")
        return None

def export_data_to_csv(df, client_name, csv_file_path):
    try:
        # Get the current date and time
        current_date = datetime.now().strftime("%Y-%m-%d")

        # Check if the DataFrame is empty
        if df.empty:
            print(f'There were no refuels on {current_date} for {client_name}.')
            return None

        # Define the CSV file name with client name and date
        csv_file_name = f'{client_name}_Report_{current_date}.csv'

        # Combine the file path and file name
        csv_file_path_string = os.path.join(csv_file_path, csv_file_name)

        # Export the DataFrame to a CSV file with headers
        df.to_csv(csv_file_path_string, index=False)

        return csv_file_path_string  # Return the file path where the CSV file was saved

    except Exception as e:
        print(f"Error exporting data to CSV: {str(e)}")
        return None